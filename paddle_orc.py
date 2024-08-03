# from src.app import app
# from src.utility.os_utils import get_file_name, join_paths
import os

import cv2
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from paddleocr import PPStructure


def join_paths(*args: str) -> str:
    """
    ### Join multiple paths together.

    This function takes a variable number of path arguments and returns a single string representing the joined path.

    #### Args:
        *args: Variable number of path arguments

    #### Returns:
        str: The joined path

    #### Example:
        >>> join_paths('path', 'to', 'file.txt')
        'path/to/file.txt'
    """
    return os.path.join(*args)


def get_file_name(path: str) -> tuple[str, str]:
    """
    ### Get the file name from a path.

    This function takes a file path and returns the file name without the extension.

    #### Args:
        path: The file path

    #### Returns:
        tuple: A tuple containing the file name and extension

    #### Example:
        >>> get_file_name('path/to/file.txt')
        ('file', '.txt')
    """
    return os.path.splitext(os.path.basename(path))


# TEMP_PATH = app.config.get("TEMP_FOLDER")
TEMP_PATH = "./temp"
# Initialize PPStructure for table extraction with recovery and OCR results
table_engine = PPStructure(
    recovery=True, return_ocr_result_in_table=True, use_gpu=False
)


# Create and save an Excel workbook to store the results
def create_workbook(workbook_name):
    WORKBOOK_PATH = join_paths(TEMP_PATH, f"{workbook_name}.xlsx")
    Workbook().save(WORKBOOK_PATH)
    book = load_workbook(WORKBOOK_PATH)
    writer = pd.ExcelWriter(WORKBOOK_PATH, engine="openpyxl", mode="a")
    writer._book = book
    writer._sheets = {ws.title: ws for ws in book.worksheets}
    return writer, book


def save_image_to_excel(image_path):
    img = cv2.imread(image_path)
    result = table_engine(img)
    image_name, _ = get_file_name(image_path)

    print("Image name %s" % image_name)

    writer, book = create_workbook(image_name)

    # Create an image object for openpyxl
    xlimg = XLImage(image_path)

    i = 1
    for line in result:
        # print("Line %d in result is: %s" % (i, line))
        # Remove the 'img' key from the result
        line.pop("img")
        # Check if the line is a table
        if line.get("type") == "table":
            # Extract HTML table and convert to DataFrame
            html_table = line.get("res").get("html")
            html_data = pd.read_html(html_table)
            df = pd.DataFrame(html_data[0])

            # Write DataFrame to Excel and add the image to the sheet
            df.to_excel(writer, sheet_name=f"table {i}", index=1)
            book[f"table {i}"].add_image(xlimg, "A100")
            i += 1

    # Save the Excel workbook
    writer.close()
    return result


# from src.utility.paddle_ocr import save_image_to_excel
image_path = "./MM-MP1044_5.png"
# image_path = "./CM-HU1157_17.png"
data = save_image_to_excel(image_path)

print("\n\n")
print(data)
print("\n\n")

for line in data:
    # line.pop("img")
    if line.get("type") == "table":
        print("Table:")
        print("\n\n")
        print(line.get("res").get("html"))
